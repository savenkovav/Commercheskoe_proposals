from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.config import (
    COMPANY_ADDRESS,
    COMPANY_DIRECTOR,
    COMPANY_INN,
    COMPANY_KPP,
    COMPANY_NAME,
    COMPANY_OGRN,
    DELIVERY_DAYS,
    DELIVERY_TERMS,
    KP_SHEET_NAME,
    KP_VAT_LABEL,
    PAYMENT_TERMS,
    WEB_PRICE_DISCOUNT_PERCENT,
)
from src.services.kp_preferences import KpPreferences, competitor_link_urls, filter_comparison_quotes
from src.services.markup_settings import get_markup_percent
from src.services.pricing_rules import (
    effective_markup_percent,
    format_markup_percent,
    pricing_adjustment_label,
    uses_web_discount_pricing,
)
from src.services.models import KitComponentLine, MatchResult, MatchSource, MatchStatus, ProposalSummary
from src.services.pricing_rules import is_internet_sourced_result
from src.services.web_quote_priority import parse_source_detail, resolve_price_source_url

STATUS_LABELS = {
    MatchStatus.EXACT: "Точное совпадение",
    MatchStatus.SIMILAR: "Похожее (проверить)",
    MatchStatus.NOT_FOUND: "Не найдено",
}

STATUS_COLORS = {
    MatchStatus.EXACT: "C6EFCE",
    MatchStatus.SIMILAR: "FFEB9C",
    MatchStatus.NOT_FOUND: "FFC7CE",
}

SOURCE_LABELS = {
    "catalog": "Каталог",
    "registry": "Реестр остатков",
    "price_list": "Прайс поставщика",
    "web": "Интернет (оценка AI)",
    "ai": "AI-подбор",
    "none": "—",
}


def _money(value: float | None) -> float:
    return round(float(value or 0), 2)


def _thin_border() -> Border:
    side = Side(style="thin", color="CCCCCC")
    return Border(left=side, right=side, top=side, bottom=side)


def _write_hyperlink(
    ws,
    row: int,
    col: int,
    url: str | None,
    display: str | None = None,
) -> None:
    label = (display or url or "—").strip()
    cell = ws.cell(row=row, column=col, value=label)
    cell.border = _thin_border()
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    if url:
        cell.hyperlink = url
        cell.font = Font(color="0563C1", underline="single")


def _specs_preview(text: str, limit: int = 500) -> str:
    cleaned = " ".join((text or "").split())
    if len(cleaned) <= limit:
        return cleaned
    return cleaned[: limit - 1] + "…"


def _margin_tender_unit(result: MatchResult) -> float:
    if result.unit_price is not None:
        return _money(result.unit_price)
    base = (
        result.unit_base_price
        if result.unit_base_price is not None
        else result.unit_cost
    )
    if base is None:
        return 0.0
    markup_pct = effective_markup_percent(result)
    if markup_pct is None:
        return 0.0
    return _money(base * (1 + markup_pct / 100))


def _margin_supplier_unit(result: MatchResult) -> float:
    if result.unit_cost is not None:
        return _money(result.unit_cost)
    if result.unit_base_price is not None:
        return _money(result.unit_base_price)
    return 0.0


class ExcelGenerator:
    def generate(
        self,
        results: list[MatchResult],
        summary: ProposalSummary,
        output_path: Path,
        request_number: str = "б/н",
        preferences: KpPreferences | None = None,
        *,
        task_mode: str = "task1",
        with_margin: bool | None = None,
        template_mode: bool = False,
    ) -> Path:
        include_links = task_mode == "task1_task2"
        show_margin = with_margin if with_margin is not None else True
        wb = Workbook()

        self._build_main_kp_sheet(
            wb.active,
            results,
            request_number,
            include_links=include_links,
            with_margin=show_margin,
        )
        if include_links and not template_mode:
            self._build_kp_specs_links_sheet(
                wb.create_sheet("Характеристики и ссылки"),
                results,
                preferences,
            )
        if show_margin and not template_mode:
            self._build_detail_sheet(
                wb.create_sheet("Детализация"), results, summary, preferences
            )
            self._build_kit_sheet(wb.create_sheet("Состав комплектов"), results)
            self._build_comparison_sheet(wb.create_sheet("Сравнение"), results)
            self._build_summary_sheet(wb.create_sheet("Сводка"), results, summary)
            self._build_margin_sheet(
                wb.create_sheet("Маржинальность"),
                results,
                preferences,
            )

        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        return output_path

    def _build_main_kp_sheet(
        self,
        ws,
        results: list[MatchResult],
        request_number: str,
        *,
        include_links: bool,
        with_margin: bool,
    ) -> None:
        ws.title = KP_SHEET_NAME
        today = datetime.now().strftime("%d.%m.%Y")
        last_col = 11 if include_links and with_margin else (10 if with_margin else 6)
        header_row = 6
        first_data_row = header_row + 1

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
        ws["A1"] = "КОММЕРЧЕСКОЕ ПРЕДЛОЖЕНИЕ"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
        ws["A2"] = f"на запрос {request_number} от {today} г."
        ws["A2"].alignment = Alignment(horizontal="center")

        legal_parts = [COMPANY_NAME]
        if COMPANY_INN:
            legal_parts.append(f"ИНН {COMPANY_INN}")
        if COMPANY_KPP:
            legal_parts.append(f"КПП {COMPANY_KPP}")
        if COMPANY_OGRN:
            legal_parts.append(f"ОГРН {COMPANY_OGRN}")
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=last_col)
        ws["A3"] = ". ".join(legal_parts)

        if COMPANY_ADDRESS:
            ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=last_col)
            ws["A4"] = COMPANY_ADDRESS

        if with_margin:
            markup_cell = ws.cell(row=5, column=8, value=get_markup_percent())
            markup_cell.font = Font(bold=True)
            markup_cell.alignment = Alignment(horizontal="center")

        headers = [
            "№ п/п",
            "Наименование товара",
            "Ед. изм.",
            "Количество",
            f"Цена, {KP_VAT_LABEL}",
            f"Стоимость, {KP_VAT_LABEL}",
        ]
        if with_margin:
            headers.extend(["%", "цена пост", "стои пост", "Поставщик"])
        if include_links and with_margin:
            headers.append("Ссылка")

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
            cell.fill = PatternFill("solid", fgColor="D9E1F2")
            cell.border = _thin_border()

        last_data_row = first_data_row - 1
        for idx, result in enumerate(results, start=first_data_row):
            last_data_row = idx
            qty = result.tz_item.quantity
            name = result.tz_item.name
            if result.status == MatchStatus.SIMILAR:
                name = f"{name} *"
            elif result.status == MatchStatus.NOT_FOUND:
                name = f"{name} (не подобрано)"

            ws.cell(row=idx, column=1, value=result.tz_item.number).border = _thin_border()
            name_cell = ws.cell(row=idx, column=2, value=name)
            name_cell.border = _thin_border()
            name_cell.alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(row=idx, column=3, value=result.tz_item.unit).border = _thin_border()
            ws.cell(row=idx, column=4, value=qty).border = _thin_border()

            unit_cost = _margin_supplier_unit(result)
            markup_pct = effective_markup_percent(result)
            if markup_pct is None:
                markup_pct = get_markup_percent()

            if with_margin:
                markup_cell = ws.cell(row=idx, column=7, value=markup_pct)
                markup_cell.number_format = "0.##"
                markup_cell.border = _thin_border()

                cost_cell = ws.cell(row=idx, column=8, value=unit_cost if unit_cost else 0)
                cost_cell.number_format = "#,##0.00"
                cost_cell.border = _thin_border()

                unit_price = _money(result.unit_price) if result.unit_price is not None else None
                line_total = _money(result.total_price) if result.total_price is not None else None
                price_cell = ws.cell(row=idx, column=5, value=unit_price if unit_price is not None else "—")
                price_cell.number_format = "#,##0.00"
                price_cell.border = _thin_border()

                total_cell = ws.cell(row=idx, column=6, value=line_total if line_total is not None else "—")
                total_cell.number_format = "#,##0.00"
                total_cell.border = _thin_border()

                supplier_total = ws.cell(row=idx, column=9, value=f"=D{idx}*H{idx}")
                supplier_total.number_format = "#,##0.00"
                supplier_total.border = _thin_border()

                supplier_name = (result.supplier or "").strip() or "—"
                if not supplier_name or supplier_name == "—":
                    supplier_name = (result.matched_name or "—")[:80]
                supplier_cell = ws.cell(row=idx, column=10, value=supplier_name)
                supplier_cell.border = _thin_border()
                supplier_cell.alignment = Alignment(wrap_text=True, vertical="top")

                if include_links and is_internet_sourced_result(result):
                    link_url = resolve_price_source_url(
                        result.comparison,
                        unit_base_price=result.unit_base_price,
                    )
                    if not link_url:
                        _, link_url = parse_source_detail(result.source_detail or "")
                    if link_url:
                        _write_hyperlink(ws, idx, 11, link_url, link_url[:80])
                    else:
                        ws.cell(row=idx, column=11, value="—").border = _thin_border()
                elif include_links:
                    ws.cell(row=idx, column=11, value="—").border = _thin_border()
            else:
                unit_price = _money(result.unit_price) if result.unit_price else None
                line_total = _money(result.total_price) if result.total_price else None
                price_cell = ws.cell(row=idx, column=5, value=unit_price or "—")
                price_cell.number_format = "#,##0.00"
                price_cell.border = _thin_border()
                total_cell = ws.cell(row=idx, column=6, value=line_total or "—")
                total_cell.number_format = "#,##0.00"
                total_cell.border = _thin_border()

        total_row = max(last_data_row + 1, first_data_row)
        ws.cell(row=total_row, column=1, value="Всего:").font = Font(bold=True)
        total_price_cell = ws.cell(row=total_row, column=6, value=f"=SUM(F{first_data_row}:F{last_data_row})")
        total_price_cell.font = Font(bold=True)
        total_price_cell.number_format = "#,##0.00"
        total_price_cell.border = _thin_border()

        if with_margin:
            supplier_sum_cell = ws.cell(
                row=total_row,
                column=9,
                value=f"=SUM(I{first_data_row}:I{last_data_row})",
            )
            supplier_sum_cell.font = Font(bold=True)
            supplier_sum_cell.number_format = "#,##0.00"
            supplier_sum_cell.border = _thin_border()

        terms_row = total_row + 2
        ws.cell(row=terms_row, column=2, value=f"Условия оплаты: {PAYMENT_TERMS}")
        ws.cell(row=terms_row + 1, column=2, value=f"Срок поставки: {DELIVERY_DAYS}")
        ws.cell(row=terms_row + 2, column=2, value=f"Доставка: {DELIVERY_TERMS}")

        if with_margin:
            tax_row = terms_row + 1
            profit_row = terms_row + 2
            margin_row = terms_row + 3
            ws.cell(row=tax_row, column=8, value="Налог").font = Font(bold=True)
            ws.cell(
                row=tax_row,
                column=9,
                value=f"=(F{total_row}-I{total_row})*5%+F{total_row}/1.05*0.05",
            ).number_format = "#,##0.00"
            ws.cell(row=profit_row, column=8, value="Прибыль").font = Font(bold=True)
            ws.cell(
                row=profit_row,
                column=9,
                value=f"=F{total_row}-I{tax_row}-I{total_row}",
            ).number_format = "#,##0.00"
            ws.cell(row=margin_row, column=8, value="Маржа").font = Font(bold=True)
            ws.cell(
                row=margin_row,
                column=9,
                value=f"=I{profit_row}/(I{total_row}+I{tax_row})",
            ).number_format = "0.00%"

        sign_row = terms_row + (5 if with_margin else 3)
        ws.cell(row=sign_row, column=1, value=COMPANY_DIRECTOR)

        note_row = sign_row + 1
        ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=last_col)
        ws.cell(
            row=note_row,
            column=1,
            value="* — позиции, требующие проверки менеджером",
        )

        widths = [6, 42, 10, 12, 18, 18]
        if with_margin:
            widths.extend([8, 14, 14, 18])
        if include_links and with_margin:
            widths.append(28)
        for col_idx, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    def _build_kp_sheet(
        self,
        ws,
        results: list[MatchResult],
        request_number: str,
    ) -> None:
        self._build_main_kp_sheet(
            ws,
            results,
            request_number,
            include_links=False,
            with_margin=False,
        )

    def _build_kp_specs_links_sheet(
        self,
        ws,
        results: list[MatchResult],
        preferences: KpPreferences | None = None,
    ) -> None:
        ws.title = "Характеристики и ссылки"
        last_col = 6

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
        ws["A1"] = "ХАРАКТЕРИСТИКИ И ССЫЛКИ К КП"
        ws["A1"].font = Font(bold=True, size=12)
        ws["A1"].alignment = Alignment(horizontal="center")

        headers = [
            "№",
            "Наименование товара",
            "Характеристики (из ТЗ)",
            "Ссылка 1",
            "Ссылка 2",
            "Ссылка 3",
        ]
        header_row = 3
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.fill = PatternFill("solid", fgColor="D9E1F2")
            cell.border = _thin_border()

        for idx, result in enumerate(results, start=1):
            row = header_row + idx
            name = result.tz_item.name
            if result.status == MatchStatus.SIMILAR:
                name = f"{name} *"
            elif result.status == MatchStatus.NOT_FOUND:
                name = f"{name} (не подобрано)"

            values = [
                result.tz_item.number,
                name,
                _specs_preview(result.tz_item.specifications, limit=2000),
            ]
            for col, value in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = _thin_border()
                cell.alignment = Alignment(wrap_text=True, vertical="top")

            links = competitor_link_urls(
                result.comparison,
                result.tz_item.name,
                preferences,
                limit=3,
            )
            for offset, url in enumerate(links):
                _write_hyperlink(ws, row, 4 + offset, url)
            for col in range(4 + len(links), 7):
                cell = ws.cell(row=row, column=col, value="—")
                cell.border = _thin_border()

        widths = [6, 36, 48, 28, 28, 28]
        for col_idx, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    def _build_detail_sheet(
        self,
        ws,
        results: list[MatchResult],
        summary: ProposalSummary,
        preferences: KpPreferences | None = None,
    ) -> None:
        headers = [
            "№",
            "Тип",
            "Запрос заказчика",
            "Найденная позиция",
            "Статус",
            "Источник",
            "Совпадение %",
            "Кол-во",
            "Себест. ед.",
            "Цена баз.",
            f"Наценка",
            "Цена ед.",
            "Сумма",
            "Поставщик",
            "Дата покупки",
            "Примечание",
            "Детали источника",
            "Альтернативы",
            "Ссылка конкурента",
        ]

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9E1F2")
            cell.border = _thin_border()
            cell.alignment = Alignment(wrap_text=True, horizontal="center")

        row = 2
        for result in results:
            row = self._write_detail_row(
                ws, row, result, row_type="Позиция ТЗ", preferences=preferences
            )
            if result.is_kit and result.kit_components:
                for comp_idx, component in enumerate(result.kit_components, start=1):
                    row = self._write_kit_component_row(
                        ws,
                        row,
                        result,
                        component,
                        comp_idx,
                        preferences=preferences,
                    )

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18
        ws.column_dimensions["C"].width = 35
        ws.column_dimensions["D"].width = 35
        ws.column_dimensions["P"].width = 40
        ws.column_dimensions["S"].width = 36

    def _write_detail_row(
        self,
        ws,
        row: int,
        result: MatchResult,
        row_type: str,
        component: KitComponentLine | None = None,
        component_index: int | None = None,
        preferences: KpPreferences | None = None,
    ) -> int:
        markup_amount = None
        markup_pct = effective_markup_percent(result) if component is None else None
        if result.unit_base_price is not None and result.unit_price is not None:
            markup_amount = _money(result.unit_price - result.unit_base_price)

        number = result.tz_item.number
        if component_index is not None:
            number = f"{result.tz_item.number}.{component_index}"

        name = result.tz_item.name
        matched_name = result.matched_name or "—"
        unit_cost = result.unit_cost
        unit_base_price = result.unit_base_price
        unit_price = result.unit_price
        total_price = result.total_price
        supplier = result.supplier
        purchase_date = result.purchase_date
        notes = result.notes
        quantity = result.tz_item.quantity
        competitor_url: str | None = None

        if component is not None:
            name = f"  ↳ {component.name}"
            matched_name = (
                component.catalog_matched_name
                if component.found_in_catalog and component.catalog_matched_name
                else component.name
            )
            unit_cost = component.unit_cost
            unit_base_price = component.unit_price
            unit_price = None
            total_price = (
                round(component.unit_price * component.quantity, 2)
                if component.unit_price is not None
                else None
            )
            supplier = component.supplier if component.found_in_catalog else None
            purchase_date = (
                component.purchase_date if component.found_in_catalog else None
            )
            notes_parts: list[str] = []
            if component.found_in_catalog:
                notes_parts.append("найдено в каталоге")
            elif component.price_list_price is not None:
                notes_parts.append(f"прайс: {component.price_list_price}")
            notes = " | ".join(notes_parts)
            quantity = component.quantity
            competitor_url = component.competitor_url
            if component.competitor_platform:
                notes = (
                    f"{notes} | {component.competitor_platform}".strip(" |")
                    if notes
                    else component.competitor_platform
                )
        else:
            links = competitor_link_urls(
                result.comparison,
                result.tz_item.name,
                preferences,
                limit=1,
            )
            competitor_url = links[0] if links else None
            if uses_web_discount_pricing(result):
                notes = (
                    f"{notes} | {pricing_adjustment_label(result)}"
                    if notes
                    else pricing_adjustment_label(result)
                )

        values = [
            number,
            row_type,
            name,
            matched_name,
            STATUS_LABELS[result.status] if component is None else "Состав комплекта",
            SOURCE_LABELS.get(result.source.value, result.source.value),
            round(result.match_score, 1) if component is None else None,
            quantity,
            unit_cost,
            unit_base_price,
            format_markup_percent(markup_pct) if markup_pct is not None else markup_amount,
            unit_price,
            total_price,
            supplier,
            purchase_date,
            notes,
            result.source_detail if component is None else "",
            "; ".join(result.alternatives[:3]) if component is None else "",
            competitor_url or "—",
        ]

        link_col = len(values)
        source_detail_col = 17
        source_detail_text = result.source_detail if component is None else ""
        source_detail_label, source_detail_url = parse_source_detail(source_detail_text)
        if component is None and not source_detail_url and is_internet_sourced_result(result):
            source_detail_url = resolve_price_source_url(
                result.comparison,
                unit_base_price=result.unit_base_price,
            )

        for col, value in enumerate(values[:-1], start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = _thin_border()
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if col in (9, 10, 11, 12, 13) and isinstance(value, (int, float)):
                cell.number_format = '#,##0.00'
            if component is not None:
                cell.fill = PatternFill("solid", fgColor="F3F4F6")

        if component is None and source_detail_url:
            _write_hyperlink(ws, row, source_detail_col, source_detail_url, source_detail_label or source_detail_text)
        elif component is None:
            ws.cell(row=row, column=source_detail_col, value=source_detail_text).border = _thin_border()

        if competitor_url:
            _write_hyperlink(ws, row, link_col, competitor_url)
        else:
            cell = ws.cell(row=row, column=link_col, value="—")
            cell.border = _thin_border()
            if component is not None:
                cell.fill = PatternFill("solid", fgColor="F3F4F6")

        if component is None:
            status_cell = ws.cell(row=row, column=5)
            status_cell.fill = PatternFill(
                "solid", fgColor=STATUS_COLORS[result.status]
            )

        return row + 1

    def _write_kit_component_row(
        self,
        ws,
        row: int,
        result: MatchResult,
        component: KitComponentLine,
        component_index: int,
        preferences: KpPreferences | None = None,
    ) -> int:
        return self._write_detail_row(
            ws,
            row,
            result,
            row_type="Составляющая",
            component=component,
            component_index=component_index,
            preferences=preferences,
        )

    def _build_kit_sheet(self, ws, results: list[MatchResult]) -> None:
        headers = [
            "№ ТЗ",
            "Комплект (позиция ТЗ)",
            "№ в составе",
            "Составляющая",
            "Кол-во",
            "Себест. ед.",
            "Цена ед.",
            "Сумма",
            "Поставщик",
            "Дата покупки",
            "Цена в прайсе",
            "Итого комплект (база)",
            "Итого комплект (КП)",
        ]

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9E1F2")
            cell.border = _thin_border()
            cell.alignment = Alignment(wrap_text=True, horizontal="center")

        row = 2
        kit_results = [r for r in results if r.is_kit and r.kit_components]
        if not kit_results:
            ws.cell(row=2, column=1, value="— комплекты с составом не найдены")
            return

        for result in kit_results:
            kit_base = result.unit_base_price
            kit_kp = result.unit_price
            first_component_row = row

            block_start = row
            for comp_idx, component in enumerate(result.kit_components, start=1):
                line_sum = (
                    round(component.unit_price * component.quantity, 2)
                    if component.unit_price is not None
                    else None
                )
                comp_supplier = component.supplier if component.found_in_catalog else None
                comp_purchase_date = (
                    component.purchase_date if component.found_in_catalog else None
                )
                display_name = component.name
                if component.found_in_catalog and component.catalog_matched_name:
                    display_name = (
                        f"{component.name} (каталог: {component.catalog_matched_name})"
                    )
                values = [
                    result.tz_item.number if comp_idx == 1 else "",
                    result.tz_item.name if comp_idx == 1 else "",
                    comp_idx,
                    display_name,
                    component.quantity,
                    component.unit_cost,
                    component.unit_price,
                    line_sum,
                    comp_supplier or "",
                    comp_purchase_date or "",
                    component.price_list_price,
                    kit_base if comp_idx == 1 else "",
                    kit_kp if comp_idx == 1 else "",
                ]
                for col, value in enumerate(values, start=1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = _thin_border()
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                    if col in (6, 7, 8, 11, 12, 13) and isinstance(value, (int, float)):
                        cell.number_format = '#,##0.00'
                row += 1

                if component.found_in_catalog and component.supplier:
                    supplier_values = [
                        "",
                        "",
                        "",
                        f"  ↳ {component.supplier}",
                        component.quantity,
                        "",
                        component.unit_price,
                        line_sum,
                        component.supplier,
                        component.purchase_date,
                        component.price_list_price,
                        "",
                        "",
                    ]
                    for col, value in enumerate(supplier_values, start=1):
                        cell = ws.cell(row=row, column=col, value=value)
                        cell.border = _thin_border()
                        cell.alignment = Alignment(wrap_text=True, vertical="top")
                        cell.fill = PatternFill("solid", fgColor="F3F4F6")
                        if col in (7, 8, 11) and isinstance(value, (int, float)):
                            cell.number_format = '#,##0.00'
                    row += 1

            if row - 1 > block_start:
                for merge_col in (1, 2, 12, 13):
                    ws.merge_cells(
                        start_row=block_start,
                        start_column=merge_col,
                        end_row=row - 1,
                        end_column=merge_col,
                    )
                    ws.cell(row=block_start, column=merge_col).alignment = Alignment(
                        vertical="top", wrap_text=True
                    )

            row += 1

        ws.column_dimensions["B"].width = 36
        ws.column_dimensions["D"].width = 42
        for col_letter in ("F", "G", "H", "K", "L", "M"):
            ws.column_dimensions[col_letter].width = 14

    def _build_comparison_sheet(self, ws, results: list[MatchResult]) -> None:
        headers = [
            "№ ТЗ",
            "Позиция ТЗ",
            "Источник",
            "Найдено",
            "Себест.",
            "Цена",
            "Поставщик",
            "Дата покупки",
            "Совпадение %",
            "Примечание",
        ]

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9E1F2")
            cell.border = _thin_border()

        row = 2
        for result in results:
            quotes = filter_comparison_quotes(list(result.comparison), KpPreferences())
            if not quotes and not result.kit_components:
                continue

            for quote in quotes:
                values = [
                    result.tz_item.number,
                    result.tz_item.name,
                    quote.label,
                    quote.matched_name,
                    quote.cost,
                    quote.price,
                    quote.supplier,
                    quote.purchase_date,
                    round(quote.match_score, 1) if quote.match_score else None,
                    quote.notes,
                ]
                for col, value in enumerate(values, start=1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = _thin_border()
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                    if col in (5, 6) and isinstance(value, (int, float)):
                        cell.number_format = '#,##0.00'
                row += 1

            for component in result.kit_components:
                values = [
                    result.tz_item.number,
                    result.tz_item.name,
                    "Состав комплекта",
                    component.name,
                    component.unit_cost,
                    component.unit_price,
                    component.supplier if component.found_in_catalog else None,
                    component.purchase_date if component.found_in_catalog else None,
                    component.quantity,
                    (
                        "каталог"
                        if component.found_in_catalog
                        else (
                            f"прайс: {component.price_list_price}"
                            if component.price_list_price is not None
                            else ""
                        )
                    ),
                ]
                for col, value in enumerate(values, start=1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = _thin_border()
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                    if col in (5, 6) and isinstance(value, (int, float)):
                        cell.number_format = '#,##0.00'
                row += 1

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 16
        ws.column_dimensions["B"].width = 32
        ws.column_dimensions["D"].width = 36

    def _build_summary_sheet(
        self, ws, results: list[MatchResult], summary: ProposalSummary
    ) -> None:
        ws["A1"] = "Сводка обработки ТЗ"
        ws["A1"].font = Font(bold=True, size=13)

        rows = [
            ("Всего позиций в ТЗ", summary.total_items),
            ("Точных совпадений", summary.exact_count),
            ("Похожих (требуют проверки)", summary.similar_count),
            ("Не найдено", summary.not_found_count),
            ("", ""),
            ("Итого себестоимость", _money(summary.total_cost)),
            ("Итого цена без наценки", _money(summary.total_base_price)),
            (
                "Наценка",
                (
                    f"{format_markup_percent(get_markup_percent())} каталог/прайс; "
                    f"{format_markup_percent(-WEB_PRICE_DISCOUNT_PERCENT)} интернет"
                ),
            ),
            ("Итого цена КП", _money(summary.total_price)),
            ("", ""),
            ("Время обработки, сек", round(summary.processing_seconds, 2)),
        ]

        for idx, (label, value) in enumerate(rows, start=3):
            ws.cell(row=idx, column=1, value=label)
            cell = ws.cell(row=idx, column=2, value=value)
            if isinstance(value, float) and label.startswith("Итого"):
                cell.number_format = '#,##0.00'

        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 20

        ws["A15"] = "Позиции без подбора:"
        ws["A15"].font = Font(bold=True)
        not_found = [r for r in results if r.status == MatchStatus.NOT_FOUND]
        if not_found:
            for idx, result in enumerate(not_found, start=16):
                ws.cell(row=idx, column=1, value=f"{result.tz_item.number}. {result.tz_item.name}")
        else:
            ws["A16"] = "— все позиции подобраны"

    def _build_margin_sheet(
        self,
        ws,
        results: list[MatchResult],
        preferences: KpPreferences | None = None,
    ) -> None:
        del preferences
        headers = [
            "№",
            "Наим. Заказчика",
            "Ед. изм.",
            "Кол-во в тенд",
            "Цена в тенд",
            "Наценка %",
            "Стоимость тенд",
            "Наим. Пост.",
            "Кол-во \nпост",
            "Цена пост",
            "Стоимость пост",
            "Маржа",
            "Прибыль/убыток",
            "Поставщик (Магазин)",
            "Ссылка",
        ]
        header_row = 1

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9E1F2")
            cell.border = _thin_border()
            cell.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")

        first_data_row = 2
        for idx, result in enumerate(results, start=first_data_row):
            qty = result.tz_item.quantity
            unit_price = _margin_tender_unit(result)
            unit_cost = _margin_supplier_unit(result)
            markup_pct = effective_markup_percent(result)
            supplier_name = (result.matched_name or "").strip() or "—"
            supplier_shop = (result.supplier or "").strip() or "—"

            link_url = None
            if is_internet_sourced_result(result):
                link_url = resolve_price_source_url(
                    result.comparison,
                    unit_base_price=result.unit_base_price,
                )
                if not link_url:
                    _, link_url = parse_source_detail(result.source_detail or "")

            ws.cell(row=idx, column=1, value=result.tz_item.number).border = _thin_border()
            ws.cell(row=idx, column=2, value=result.tz_item.name).border = _thin_border()
            ws.cell(row=idx, column=3, value=result.tz_item.unit).border = _thin_border()
            ws.cell(row=idx, column=4, value=qty).border = _thin_border()

            price_cell = ws.cell(row=idx, column=5, value=unit_price)
            price_cell.number_format = "#,##0.00"
            price_cell.border = _thin_border()

            markup_cell = ws.cell(
                row=idx,
                column=6,
                value=format_markup_percent(markup_pct),
            )
            markup_cell.border = _thin_border()
            markup_cell.alignment = Alignment(horizontal="center", vertical="top")

            cost_cell = ws.cell(row=idx, column=10, value=unit_cost)
            cost_cell.number_format = "#,##0.00"
            cost_cell.border = _thin_border()

            ws.cell(row=idx, column=8, value=supplier_name).border = _thin_border()
            ws.cell(row=idx, column=14, value=supplier_shop).border = _thin_border()

            ws.cell(row=idx, column=7, value=f"=D{idx}*E{idx}").number_format = "#,##0.00"
            ws.cell(row=idx, column=7).border = _thin_border()
            ws.cell(row=idx, column=9, value=f"=D{idx}").border = _thin_border()
            ws.cell(row=idx, column=11, value=f"=I{idx}*J{idx}").number_format = '#,##0.00 "₽"'
            ws.cell(row=idx, column=11).border = _thin_border()
            ws.cell(row=idx, column=12, value=f"=IF(G{idx}=0,0,(G{idx}-K{idx})/G{idx})").number_format = "0%"
            ws.cell(row=idx, column=12).border = _thin_border()
            ws.cell(row=idx, column=13, value=f"=(G{idx}-K{idx})").number_format = "#,##0.00"
            ws.cell(row=idx, column=13).border = _thin_border()

            if link_url:
                _write_hyperlink(ws, idx, 15, link_url)
            else:
                cell = ws.cell(row=idx, column=15, value="—")
                cell.border = _thin_border()

            for col in (2, 3, 4, 8, 14):
                ws.cell(row=idx, column=col).alignment = Alignment(
                    wrap_text=True,
                    vertical="top",
                )

        if not results:
            ws.cell(row=2, column=2, value="— позиции не найдены")
            last_data_row = 1
        else:
            last_data_row = first_data_row + len(results) - 1

        summary_row = last_data_row + 1
        expense_included_row = summary_row + 1
        expense_excluded_row = summary_row + 2
        tax_row = summary_row + 3
        total_cost_row = summary_row + 4
        margin_row = summary_row + 5
        roi_row = summary_row + 6

        def _label(row: int, text: str, *, bold: bool = True) -> None:
            cell = ws.cell(row=row, column=2, value=text)
            cell.font = Font(bold=bold)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        def _money_formula(row: int, col: int, formula: str) -> None:
            cell = ws.cell(row=row, column=col, value=formula)
            cell.number_format = "#,##0.00"
            cell.font = Font(bold=True)
            cell.border = _thin_border()

        if last_data_row >= first_data_row:
            sum_range_g = f"G{first_data_row}:G{last_data_row}"
            sum_range_k = f"K{first_data_row}:K{last_data_row}"
        else:
            sum_range_g = "G2:G2"
            sum_range_k = "K2:K2"

        _label(summary_row, "цена контракта")
        ws.cell(row=summary_row, column=10, value="затраты").font = Font(bold=True)
        _money_formula(summary_row, 7, f"=SUM({sum_range_g})")
        _money_formula(summary_row, 11, f"=SUM({sum_range_k})")

        _label(expense_included_row, "учитывается в расходах")
        _money_formula(
            expense_included_row,
            11,
            f"=K{summary_row}-K{expense_excluded_row}",
        )

        _label(expense_excluded_row, "не учитывается в расходах")
        excluded_cell = ws.cell(row=expense_excluded_row, column=11, value=0)
        excluded_cell.font = Font(bold=True)
        excluded_cell.border = _thin_border()

        _label(tax_row, "Налог")
        _money_formula(
            tax_row,
            11,
            f"=G{summary_row}/1.05*0.05+(G{summary_row}-K{expense_included_row})*5%",
        )

        _label(total_cost_row, "Затраты общие")
        _money_formula(
            total_cost_row,
            11,
            f"=SUM(K{expense_included_row}:K{tax_row})",
        )

        _label(margin_row, "Маржа")
        _money_formula(margin_row, 11, f"=G{summary_row}-K{total_cost_row}")

        _label(roi_row, "Процент доходности")
        roi_cell = ws.cell(
            row=roi_row,
            column=11,
            value=f"=IF(K{total_cost_row}=0,0,K{margin_row}/K{total_cost_row}*100)",
        )
        roi_cell.number_format = "0"
        roi_cell.font = Font(bold=True)
        roi_cell.border = _thin_border()

        note_row = roi_row + 2
        ws.cell(
            row=note_row,
            column=2,
            value=(
                f"Наценка: {format_markup_percent(get_markup_percent())} для каталога/прайсов; "
                f"{format_markup_percent(-WEB_PRICE_DISCOUNT_PERCENT)} для цен из интернета"
            ),
        )
        ws.cell(row=note_row, column=2).alignment = Alignment(wrap_text=True, vertical="top")

        widths = [6, 34, 10, 12, 14, 10, 14, 28, 10, 12, 14, 10, 14, 18, 32]
        for col_idx, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width
