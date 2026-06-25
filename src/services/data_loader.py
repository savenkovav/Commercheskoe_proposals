from __future__ import annotations

import logging
import re
from pathlib import Path

import xlrd
from openpyxl import load_workbook

from src.config import REGISTRY_PHOTOS_DIR
from src.services.models import CatalogItem, GoodsReportItem, PriceListItem, RegistryItem

logger = logging.getLogger(__name__)

REGISTRY_PHOTO_COLUMNS = range(4, 7)


def _normalize_name(value: str) -> str:
    text = value.lower().strip()
    text = re.sub(r"\s+", " ", text)
    text = re.sub(r"[«»\"']", "", text)
    return text


def _classify_catalog_row(name: str, cost: float | None, price: float | None) -> tuple[str, str | None]:
    lower = name.lower()
    if "составляющие" in lower:
        return "components_header", name
    if cost is None and price is None:
        return "section", None
    if "комплект" in lower and price is not None and cost is None:
        return "kit_total", None
    if "комплект" in lower and cost is not None:
        return "sub_kit", None
    return "item", None


def _parse_optional_number(value) -> float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("\u00a0", " ")
    if not text or text in {"-", "—", "–", " " * len(text)}:
        return None
    cleaned = text.replace(" ", "").replace(",", ".")
    try:
        return float(cleaned)
    except ValueError:
        return None


def _parse_catalog_unit(value) -> str:
    if value is None:
        return "шт"
    text = str(value).strip()
    return text or "шт"


def _is_supplier_header_name(name: str) -> bool:
    lower = name.lower()
    return lower.startswith("состав") or "составляющие" in lower


def format_catalog_supplier(item: CatalogItem) -> str | None:
    parts: list[str] = []
    if item.supplier and item.supplier.strip():
        parts.append(item.supplier.strip())
    if item.supplier_note and item.supplier_note.strip():
        note = item.supplier_note.strip()
        if note not in parts:
            parts.append(note)
    if not parts:
        return None
    return "\n".join(parts)


def build_catalog_rag_text(catalog: list[CatalogItem]) -> str:
    lines = ["# Каталог товаров", ""]
    for item in catalog:
        if item.entry_type not in {"item", "kit_total", "sub_kit"}:
            continue
        supplier = format_catalog_supplier(item)
        parts = [f"Строка {item.row_index}: {item.name}"]
        if supplier:
            parts.append(f"Поставщик: {supplier.replace(chr(10), '; ')}")
        if item.cost is not None:
            parts.append(f"Себестоимость: {item.cost}")
        if item.price is not None:
            parts.append(f"Цена: {item.price}")
        if item.stock is not None:
            parts.append(f"Остаток: {item.stock} {item.unit}")
        if item.actual_markup_pct is not None:
            parts.append(f"Наценка: {item.actual_markup_pct}%")
        parts.append(f"Ед. изм.: {item.unit or 'шт'}")
        lines.append(" | ".join(parts))
    return "\n".join(lines)


def load_catalog(path: Path) -> list[CatalogItem]:
    if not path.exists():
        return []

    items: list[CatalogItem] = []
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    current_group: str | None = None
    current_supplier: str | None = None
    current_supplier_note: str | None = None

    for row_idx in range(2, ws.max_row + 1):
        cell_a = ws.cell(row_idx, 1)
        name = str(cell_a.value).strip() if cell_a.value else ""
        if not name:
            continue

        cost = _parse_optional_number(ws.cell(row_idx, 4).value)
        sale_price = _parse_optional_number(ws.cell(row_idx, 2).value)
        stock = _parse_optional_number(ws.cell(row_idx, 3).value)
        unit = _parse_catalog_unit(ws.cell(row_idx, 5).value)
        markup = _parse_optional_number(ws.cell(row_idx, 6).value)

        bold = bool(cell_a.font and cell_a.font.bold)
        is_supplier_row = (
            bold
            and cost is None
            and sale_price is None
            and stock is None
            and markup is None
        )

        if is_supplier_row:
            if _is_supplier_header_name(name):
                current_supplier_note = name
                current_group = name
                entry_type = "components_header"
            else:
                current_supplier = name
                current_supplier_note = None
                current_group = None
                entry_type = "section"
            items.append(
                CatalogItem(
                    name=name,
                    cost=None,
                    price=None,
                    unit=unit or "шт",
                    stock=None,
                    source_file=path.name,
                    actual_markup_pct=None,
                    entry_type=entry_type,
                    components_group=current_group if entry_type == "components_header" else None,
                    row_index=row_idx,
                    supplier=current_supplier,
                    supplier_note=current_supplier_note,
                )
            )
            continue

        entry_type, group_name = _classify_catalog_row(
            name,
            cost,
            sale_price,
        )
        if entry_type == "components_header":
            current_group = group_name
            current_supplier_note = name
        elif entry_type == "section":
            current_group = None
            current_supplier = name
            current_supplier_note = None

        items.append(
            CatalogItem(
                name=name,
                cost=cost,
                price=sale_price,
                unit=unit or "шт",
                stock=stock,
                source_file=path.name,
                actual_markup_pct=markup,
                entry_type=entry_type,
                components_group=current_group,
                row_index=row_idx,
                supplier=current_supplier,
                supplier_note=current_supplier_note,
            )
        )

    wb.close()
    return items


def _parse_excel_date(value) -> str | None:
    if value is None:
        return None
    if hasattr(value, "strftime"):
        return value.strftime("%d.%m.%Y")
    text = str(value).strip()
    return text or None


_SUPPLIER_MARKERS = re.compile(
    r"\b(ооо|оао|зао|пао|ип|ао|нпо|тд|чп)\b|[«\"]",
    re.IGNORECASE,
)


def _looks_like_supplier(name: str) -> bool:
    return bool(_SUPPLIER_MARKERS.search(name))


def _is_procurement_report(ws) -> bool:
    for row in ws.iter_rows(min_row=1, max_row=12, values_only=True):
        for cell in row[:3]:
            if cell and "отчет по закупкам" in str(cell).lower():
                return True
    return False


def _is_stock_balance_report(ws) -> bool:
    for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
        for cell in row[:3]:
            if cell and "остатки на дату" in str(cell).lower():
                return True
    return False


def _is_stock_balance_report_rows(rows: list[tuple]) -> bool:
    for row in rows:
        for cell in row[:3]:
            if cell and "остатки на дату" in str(cell).lower():
                return True
    return False


_TENDER_HEADER = re.compile(
    r"^(?:\d+\.\s*)?(?:тендер|тендеры)\b",
    re.IGNORECASE,
)


def _looks_like_stock_section_row(name: str) -> bool:
    lower = name.lower().strip()
    if not lower:
        return True
    if _looks_like_supplier(name):
        return True
    if _TENDER_HEADER.search(lower):
        return True
    if re.fullmatch(r"\d{2}-\d{2}(?:\.\s.*)?", lower):
        return True
    if lower.startswith("построен:"):
        return True
    return False


def _parse_stock_balance_unit(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text or len(text) > 12:
        return None
    return text


def _load_stock_balance_report(ws, source_file: str) -> list[GoodsReportItem]:
    items: list[GoodsReportItem] = []
    for row in ws.iter_rows(min_row=13, values_only=True):
        if not row:
            continue
        name = str(row[0]).strip() if row[0] else ""
        if not name or _looks_like_stock_section_row(name):
            continue

        unit = _parse_stock_balance_unit(row[3] if len(row) > 3 else None)
        cost = _parse_optional_number(row[4] if len(row) > 4 else None)
        if not unit or cost is None or cost <= 0:
            continue

        items.append(
            GoodsReportItem(
                name=name,
                supplier=None,
                purchase_date=None,
                cost=float(cost),
                price=float(cost),
                unit=unit,
                source_file=f"stock:{source_file}",
            )
        )
    return items


def _load_stock_balance_registry(ws) -> list[RegistryItem]:
    items: list[RegistryItem] = []
    for row in ws.iter_rows(min_row=13, values_only=True):
        if not row:
            continue
        name = str(row[0]).strip() if row[0] else ""
        if not name or _looks_like_stock_section_row(name):
            continue

        unit = _parse_stock_balance_unit(row[3] if len(row) > 3 else None)
        cost = _parse_optional_number(row[4] if len(row) > 4 else None)
        if not unit or cost is None or cost <= 0:
            continue

        qty = row[2] if len(row) > 2 and isinstance(row[2], (int, float)) else 0.0
        items.append(
            RegistryItem(
                name=name,
                quantity=float(qty),
                condition=None,
                link=None,
                photo_files=[],
            )
        )
    return items


def _load_standard_goods_report(ws, source_file: str) -> list[GoodsReportItem]:
    items: list[GoodsReportItem] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = str(row[0]).strip() if row[0] else ""
        if not name:
            continue

        supplier = str(row[1]).strip() if len(row) > 1 and row[1] else None
        purchase_date = _parse_excel_date(row[2]) if len(row) > 2 else None
        cost = row[3] if len(row) > 3 and isinstance(row[3], (int, float)) else None
        sale_price = row[4] if len(row) > 4 and isinstance(row[4], (int, float)) else None
        unit = str(row[5]).strip() if len(row) > 5 and row[5] else "шт"

        items.append(
            GoodsReportItem(
                name=name,
                supplier=supplier or None,
                purchase_date=purchase_date,
                cost=float(cost) if cost is not None else None,
                price=float(sale_price) if sale_price is not None else None,
                unit=unit or "шт",
                source_file=source_file,
            )
        )
    return items


def _load_procurement_report(ws, source_file: str) -> list[GoodsReportItem]:
    rows = list(ws.iter_rows(values_only=True))
    items: list[GoodsReportItem] = []
    period: str | None = None

    for row in rows[:6]:
        cell = row[0] if row else None
        if cell and re.search(r"\d{2}\.\d{2}\.\d{4}", str(cell)):
            period = str(cell).strip()

    idx = 0
    while idx < len(rows):
        row = rows[idx]
        name = str(row[0]).strip() if row and row[0] else ""
        qty = row[1] if row and len(row) > 1 else None
        unit = row[2] if row and len(row) > 2 else None
        price = row[3] if row and len(row) > 3 else None

        idx += 1
        if not name or not isinstance(qty, (int, float)) or not isinstance(price, (int, float)):
            continue
        if not unit or not str(unit).strip():
            continue

        supplier: str | None = None
        if idx < len(rows):
            next_row = rows[idx]
            next_name = str(next_row[0]).strip() if next_row and next_row[0] else ""
            next_unit = next_row[2] if next_row and len(next_row) > 2 else None
            if next_name and not (next_unit and str(next_unit).strip()) and _looks_like_supplier(
                next_name
            ):
                supplier = next_name
                idx += 1

        items.append(
            GoodsReportItem(
                name=name,
                supplier=supplier,
                purchase_date=period,
                cost=float(price),
                price=float(price),
                unit=str(unit).strip() or "шт",
                source_file=f"procurement:{source_file}",
            )
        )

    return items


def load_goods_report(path: Path) -> list[GoodsReportItem]:
    if not path.exists():
        return []

    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    if _is_stock_balance_report(ws):
        items = _load_stock_balance_report(ws, path.name)
    elif _is_procurement_report(ws):
        items = _load_procurement_report(ws, path.name)
    else:
        items = _load_standard_goods_report(ws, path.name)
    wb.close()
    return items


def merge_goods_reports(*sources: list[GoodsReportItem]) -> list[GoodsReportItem]:
    merged: list[GoodsReportItem] = []
    seen: set[str] = set()
    for items in sources:
        for item in items:
            key = normalize_name(item.name)
            if key in seen:
                continue
            seen.add(key)
            merged.append(item)
    return merged


def merge_registry(*sources: list[RegistryItem]) -> list[RegistryItem]:
    merged: list[RegistryItem] = []
    seen: set[str] = set()
    for items in sources:
        for item in items:
            key = normalize_name(item.name)
            if key in seen:
                continue
            seen.add(key)
            merged.append(item)
    return merged


def _extract_registry_photos(path: Path, photos_dir: Path) -> None:
    photos_dir.mkdir(parents=True, exist_ok=True)
    marker = photos_dir / ".extracted"

    if marker.exists() and marker.stat().st_mtime >= path.stat().st_mtime:
        return

    for cached in photos_dir.iterdir():
        if cached.name != ".extracted":
            cached.unlink()

    wb = load_workbook(path, read_only=False)
    ws = wb[wb.sheetnames[0]]
    row_images: dict[int, list[tuple[int, object]]] = {}

    for image in ws._images:
        anchor = image.anchor
        if not hasattr(anchor, "_from"):
            continue
        row_idx = anchor._from.row
        col_idx = anchor._from.col
        if col_idx in REGISTRY_PHOTO_COLUMNS:
            row_images.setdefault(row_idx, []).append((col_idx, image))

    extracted = 0
    for row_idx, images in row_images.items():
        item_idx = row_idx - 1
        if item_idx < 0:
            continue
        for photo_num, (_, image) in enumerate(sorted(images, key=lambda item: item[0])):
            suffix = Path(image.path or "photo.png").suffix or ".png"
            filename = (
                f"{item_idx:04d}{suffix.lower()}"
                if photo_num == 0
                else f"{item_idx:04d}_{photo_num + 1}{suffix.lower()}"
            )
            (photos_dir / filename).write_bytes(image._data())
            extracted += 1

    marker.touch()
    wb.close()
    logger.info("Извлечено фото из реестра: %s", extracted)


def _resolve_registry_photo_files(photos_dir: Path, item_idx: int) -> list[str]:
    files: list[str] = []
    prefixes = [f"{item_idx:04d}"] + [f"{item_idx:04d}_{num}" for num in range(2, 10)]
    for prefix in prefixes:
        found: str | None = None
        for suffix in (".png", ".jpg", ".jpeg", ".gif", ".webp"):
            candidate = f"{prefix}{suffix}"
            if (photos_dir / candidate).is_file():
                found = candidate
                break
        if not found:
            break
        files.append(found)
    return files


def load_registry(path: Path, photos_dir: Path | None = None) -> list[RegistryItem]:
    if not path.exists():
        return []

    cache_dir = photos_dir or REGISTRY_PHOTOS_DIR

    probe_wb = load_workbook(path, read_only=True, data_only=True)
    probe_ws = probe_wb[probe_wb.sheetnames[0]]
    header_rows = list(probe_ws.iter_rows(min_row=1, max_row=10, values_only=True))
    is_stock = _is_stock_balance_report_rows(header_rows)
    probe_wb.close()

    if is_stock:
        wb = load_workbook(path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        items = _load_stock_balance_registry(ws)
        wb.close()
        return items

    _extract_registry_photos(path, cache_dir)

    items: list[RegistryItem] = []
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if not row:
            continue
        name = str(row[0]).strip() if row[0] else ""
        if not name:
            continue

        quantity = row[1] if len(row) > 1 and isinstance(row[1], (int, float)) else 0.0
        condition = str(row[3]).strip() if len(row) > 3 and row[3] else None
        link = str(row[7]).strip() if len(row) > 7 and row[7] else None
        photo_files = _resolve_registry_photo_files(cache_dir, idx)

        items.append(
            RegistryItem(
                name=name,
                quantity=float(quantity),
                condition=condition if condition and condition != "None" else None,
                link=link if link and link != "None" else None,
                photo_files=photo_files,
            )
        )

    wb.close()
    return items


def _optional_number(value: object) -> float | None:
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        text = value.strip().replace(",", ".")
        if not text:
            return None
        try:
            return float(text)
        except ValueError:
            return None
    return None


def _parse_price_row(row: list, sheet_name: str, supplier_name: str) -> PriceListItem | None:
    if len(row) < 4:
        return None

    code = str(row[0]).strip() if row[0] else ""
    name = str(row[1]).strip() if row[1] else ""
    price = row[3] if isinstance(row[3], (int, float)) else None
    recommended_qty = _optional_number(row[2]) if len(row) > 2 else None
    order_qty = _optional_number(row[4]) if len(row) > 4 else None
    order_sum = _optional_number(row[5]) if len(row) > 5 else None

    if not name or price is None or price <= 0:
        return None
    if not code or len(code) > 8:
        return None

    return PriceListItem(
        code=code,
        name=name,
        price=float(price),
        sheet=sheet_name,
        supplier=supplier_name,
        recommended_qty=recommended_qty,
        order_qty=order_qty,
        order_sum=order_sum,
    )


def _load_price_list_xls(path: Path, supplier_name: str) -> list[PriceListItem]:
    items: list[PriceListItem] = []
    wb = xlrd.open_workbook(str(path))
    skip_sheets = {"Лист1", "ПРОСМОТР ЗАКАЗА"}

    for sheet_name in wb.sheet_names():
        if sheet_name in skip_sheets:
            continue

        ws = wb.sheet_by_name(sheet_name)
        for row_idx in range(ws.nrows):
            item = _parse_price_row(ws.row_values(row_idx), sheet_name, supplier_name)
            if item:
                items.append(item)

    return items


def _load_price_list_xlsx(path: Path, supplier_name: str) -> list[PriceListItem]:
    items: list[PriceListItem] = []
    wb = load_workbook(path, read_only=True, data_only=True)
    skip_sheets = {"Лист1", "ПРОСМОТР ЗАКАЗА"}

    for sheet_name in wb.sheetnames:
        if sheet_name in skip_sheets:
            continue

        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True):
            item = _parse_price_row(list(row), sheet_name, supplier_name)
            if item:
                items.append(item)

    wb.close()
    return items


def load_price_list(path: Path, supplier: str | None = None) -> list[PriceListItem]:
    if not path.exists():
        return []

    supplier_name = supplier or path.stem
    suffix = path.suffix.lower()

    if suffix == ".xls":
        return _load_price_list_xls(path, supplier_name)
    if suffix == ".xlsx":
        return _load_price_list_xlsx(path, supplier_name)

    raise ValueError(f"Неподдерживаемый формат прайса: {suffix}")


def parse_tz_docx(path: Path) -> list:
    from src.services.tz_parser import parse_tz

    return parse_tz(path)


def normalize_name(value: str) -> str:
    return _normalize_name(value)
